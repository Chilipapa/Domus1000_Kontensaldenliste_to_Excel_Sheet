import re
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from pathlib import Path
from datetime import datetime

# Pfade
docx_path = Path(__file__).parent / "d600_d010_Kontensalden.docx"
excel_path = Path(__file__).parent / "Konten_Auszug.xlsx"
filtered_excel_path = Path(__file__).parent / "Konten_Auszug_gefiltert.xlsx"

# Word-Datei laden
print(f"Lese Word-Datei: {docx_path}")
doc = Document(docx_path)

# Daten sammeln
accounts = []
account_pattern = re.compile(r'^(\d{4,5})\s+(.*)$')
year = ""

for paragraph in doc.paragraphs:
    text = paragraph.text.strip()
    if not text:
        continue
    
    # Extrahiere Jahr aus Buchungszeitraum-Zeile
    if "Buchungszeitraum:" in text:
        year_match = re.search(r'\b(\d{4})\b', text)
        if year_match:
            year = year_match.group(1)
    
    match = account_pattern.match(text)
    if match:
        account_number = match.group(1)
        rest = match.group(2)
        
        # Zahlen extrahieren
        numbers_found = re.findall(r'[\d.,]+', rest)
        cleaned_numbers = []
        for num in numbers_found:
            cleaned = num.replace('.', '')
            if cleaned:
                cleaned_numbers.append(cleaned)
        
        # Text vor Zahlen extrahieren
        text_part = re.sub(r'\s+[\d.,].*$', '', rest).strip()
        
        accounts.append({
            'Kontonummer': account_number,
            'Text': text_part,
            'Zahlen': cleaned_numbers
        })

# Funktion zum Erstellen einer Excel-Datei
def create_excel(output_path, data_accounts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Konten"
    
    # Drei Kopfzeilen (Firma / Adresse / Leerzeile)
    ws['A1'] = "Grundstücksgemeinschaft Sabine und Horst Firek"
    ws['A2'] = "Magistratsweg 126/128 13591 Berlin und Blankeneser Weg 1A/1B 13581 Berlin"
    ws['A3'] = ""

    # Header (in Zeile 4)
    headers = ["Kontonummer", "Vollständiger Text", "Ausgaben", "Einnahmen", "Saldo"]
    ws.append(headers)  # nach den drei gesetzten Zeilen landet der Header in Zeile 4

    # Header-Formatierung (Zeile 4)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Daten hinzufügen
    ausgaben_gesamt = 0
    einnahmen_gesamt = 0
    
    for account in data_accounts:
        ausgaben = account['Zahlen'][0] if len(account['Zahlen']) > 0 else ""
        einnahmen = account['Zahlen'][1] if len(account['Zahlen']) > 1 else ""
        
        # Summen berechnen
        if ausgaben:
            try:
                ausgaben_gesamt += float(ausgaben.replace(',', '.'))
            except:
                pass
        if einnahmen:
            try:
                einnahmen_gesamt += float(einnahmen.replace(',', '.'))
            except:
                pass
        
        # Row hinzufügen
        row = [
            account['Kontonummer'],
            account['Text'],
            "",
            "",
            ""
        ]
        ws.append(row)
        
        # Werte und Formatierung setzen
        row_num = ws.max_row
        if ausgaben:
            ws[f'C{row_num}'] = float(ausgaben.replace(',', '.'))
            ws[f'C{row_num}'].number_format = '#,##0.00 "€"'
        else:
            ws[f'C{row_num}'] = 0
            ws[f'C{row_num}'].number_format = '- "€"'
            
        if einnahmen:
            ws[f'D{row_num}'] = float(einnahmen.replace(',', '.'))
            ws[f'D{row_num}'].number_format = '#,##0.00 "€"'
        else:
            ws[f'D{row_num}'] = 0
            ws[f'D{row_num}'].number_format = '- "€"'
        
        # Saldo-Formel
        ws[f'E{row_num}'] = f"=D{row_num}-C{row_num}"
        ws[f'E{row_num}'].number_format = '#,##0.00 "€"'
    
    # Summen-Zeile hinzufügen (Daten beginnen in Zeile 5)
    summen_row = ws.max_row + 1
    ws[f'A{summen_row}'] = "SUMME"
    ws[f'C{summen_row}'] = f"=SUM(C5:C{summen_row-1})"
    ws[f'D{summen_row}'] = f"=SUM(D5:D{summen_row-1})"
    ws[f'E{summen_row}'] = f"=D{summen_row}-C{summen_row}"
    
    # Formatierung Summen
    ws[f'C{summen_row}'].number_format = '#,##0.00 "€"'
    ws[f'D{summen_row}'].number_format = '#,##0.00 "€"'
    ws[f'E{summen_row}'].number_format = '#,##0.00 "€"'
    
    # Summen-Zeile formatieren: keine Füllung, schwarzer fetter Text
    summen_fill = PatternFill()  # default: no fill
    summen_font = Font(bold=True, color="000000")
    for col in ['A', 'C', 'D', 'E']:
        ws[f'{col}{summen_row}'].fill = summen_fill
        ws[f'{col}{summen_row}'].font = summen_font
    
    # Spaltenbreite
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    # Speichern
    wb.save(output_path)
    
    # Rückgabe der Gesamtsummen
    return ausgaben_gesamt, einnahmen_gesamt

# Alle Daten exportieren
print(f"\nErstelle Excel-Datei: {excel_path}")
ausgaben_gesamt, einnahmen_gesamt = create_excel(excel_path, accounts)

# Gefilterte Daten exportieren (80000-80200 und 89100)
print(f"Erstelle gefilterte Excel-Datei: {filtered_excel_path}")
filtered_accounts = []
for account in accounts:
    konto_num = int(account['Kontonummer'])
    if (80000 <= konto_num <= 80200) or konto_num == 89100:
        filtered_accounts.append(account)

# Spezielle Behandlung für gefilterte Datei
wb_filtered = Workbook()
ws_filtered = wb_filtered.active
ws_filtered.title = "Konten"

# Drei Kopfzeilen (Firma / Adresse / Leerzeile) oben
ws_filtered['A1'] = "Grundstücksgemeinschaft Sabine und Horst Firek"
ws_filtered['A2'] = "Magistratsweg 126/128 13591 Berlin und Blankeneser Weg 1A/1B 13581 Berlin"
ws_filtered['A3'] = ""

# Drei neue leere Zeilen unter Zeile 3
ws_filtered['A4'] = ""
ws_filtered['A5'] = ""
ws_filtered['A6'] = ""

# Titel in Zeile 7
if not year:
    year = datetime.now().year
ws_filtered['A7'] = f"Ausgaben für BK und HK {year}"
ws_filtered['A7'].font = Font(bold=True, size=12)

# Aktuelles Datum in Zeile 8 (nur in Spalte E)
ws_filtered['A8'] = ""
ws_filtered['E8'] = datetime.now().strftime('%d.%m.%Y')
ws_filtered['E8'].font = Font(italic=True)

# Header in Zeile 9
headers = ["Kontonummer", "Vollständiger Text", "Ausgaben", "Einnahmen", "Saldo"]
ws_filtered.append([])
ws_filtered.append(headers)

# Header-Formatierung (Zeile 9) - keine Füllung, schwarzer Text
header_fill = PatternFill()  # no fill
header_font = Font(bold=True, color="000000")
for cell in ws_filtered[9]:
    cell.fill = header_fill
    cell.font = header_font

# Daten hinzufügen
ausgaben_gefiltert = 0
einnahmen_gefiltert = 0

for account in filtered_accounts:
    ausgaben = account['Zahlen'][0] if len(account['Zahlen']) > 0 else ""
    einnahmen = account['Zahlen'][1] if len(account['Zahlen']) > 1 else ""
    
    # Summen berechnen
    if ausgaben:
        try:
            ausgaben_gefiltert += float(ausgaben.replace(',', '.'))
        except:
            pass
    if einnahmen:
        try:
            einnahmen_gefiltert += float(einnahmen.replace(',', '.'))
        except:
            pass
    
    # Row hinzufügen
    row = [
        account['Kontonummer'],
        account['Text'],
        "",
        "",
        ""
    ]
    ws_filtered.append(row)
    
    # Werte und Formatierung setzen
    row_num = ws_filtered.max_row
    if ausgaben:
        ws_filtered[f'C{row_num}'] = float(ausgaben.replace(',', '.'))
        ws_filtered[f'C{row_num}'].number_format = '#,##0.00 "€"'
    else:
        ws_filtered[f'C{row_num}'] = 0
        ws_filtered[f'C{row_num}'].number_format = '- "€"'
        
    if einnahmen:
        ws_filtered[f'D{row_num}'] = float(einnahmen.replace(',', '.'))
        ws_filtered[f'D{row_num}'].number_format = '#,##0.00 "€"'
    else:
        ws_filtered[f'D{row_num}'] = 0
        ws_filtered[f'D{row_num}'].number_format = '- "€"'
    
    # Saldo-Formel
    ws_filtered[f'E{row_num}'] = f"=C{row_num}-D{row_num}"
    ws_filtered[f'E{row_num}'].number_format = '#,##0.00 "€"'

# Summen-Zeile hinzufügen
summen_row = ws_filtered.max_row + 1
ws_filtered[f'A{summen_row}'] = "SUMME BETRIEBSKOSTEN"
ws_filtered[f'C{summen_row}'] = f"=SUM(C5:C{summen_row-1})"
ws_filtered[f'D{summen_row}'] = f"=SUM(D5:D{summen_row-1})"
ws_filtered[f'E{summen_row}'] = f"=C{summen_row}-D{summen_row}"

# Formatierung Summen
ws_filtered[f'C{summen_row}'].number_format = '#,##0.00 "€"'
ws_filtered[f'D{summen_row}'].number_format = '#,##0.00 "€"'
ws_filtered[f'E{summen_row}'].number_format = '#,##0.00 "€"'

# Summen-Zeile formatieren: keine Füllung, schwarzer fetter Text
summen_fill = PatternFill()
summen_font = Font(bold=True, color="000000")
for col in ['A', 'C', 'D', 'E']:
    ws_filtered[f'{col}{summen_row}'].fill = summen_fill
    ws_filtered[f'{col}{summen_row}'].font = summen_font

# 3 Leerzeilen hinzufügen (direkt nach summen_row)
ws_filtered[f'A{summen_row + 1}'] = ""
ws_filtered[f'A{summen_row + 2}'] = ""
ws_filtered[f'A{summen_row + 3}'] = ""

# Zusätzliche Leerzeile und Erläuterungstext hinzufügen
text_start_row = summen_row + 2
ws_filtered[f'A{text_start_row}'] = ""  # Leerzeile
ws_filtered[f'A{text_start_row + 1}'] = "Wichtig für Wiso-Steuer ist"
ws_filtered[f'A{text_start_row + 2}'] = "nur der Saldo der Betriebskosten-Aufwendungen."
ws_filtered[f'A{text_start_row + 3}'] = "Warum?"
ws_filtered[f'A{text_start_row + 4}'] = "Wir buchen die Gutschriften, Umtausche oder Erstattungen"
ws_filtered[f'A{text_start_row + 5}'] = "nicht als extra Einnahme, sie vermindern nur die Aufwendungen."
ws_filtered[f'A{text_start_row + 6}'] = "Es werden, wie bei der BK-Abrechnung auch,"
ws_filtered[f'A{text_start_row + 7}'] = "die saldierten Aufwendungen, also reale Kosten, angesetzt."

# Zusätzliche Leerzeile und weiterer Erläuterungstext
ws_filtered[f'A{text_start_row + 8}'] = ""  # Leerzeile
ws_filtered[f'A{text_start_row + 9}'] = "Dieser Betrag muss bei Wiso-Steuer eingetragen werden bzw. herauskommen."
ws_filtered[f'E{text_start_row + 9}'] = f"=E{summen_row}"  # Saldo-Wert aus SUMME BETRIEBSKOSTEN
ws_filtered[f'E{text_start_row + 9}'].number_format = '#,##0.00 "€"'
ws_filtered[f'E{text_start_row + 9}'].font = Font(bold=True, color="000000")  # Fett formatieren
ws_filtered[f'E{text_start_row + 9}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Rahmen
ws_filtered[f'E{text_start_row + 9}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Hellblaue Füllung
ws_filtered[f'A{text_start_row + 10}'] = "Dieser Betrag wird ebenfalls in die Excel-Datei"
ws_filtered[f'A{text_start_row + 11}'] = "nicht_umlagefähige_kosten.xlsx, in die Zeile"
ws_filtered[f'A{text_start_row + 12}'] = "Domus1000 Ausgaben BK u. HK als MINUS-ZAHL eingetragen"
ws_filtered[f'E{text_start_row + 12}'] = f"=-E{summen_row}"  # Saldo-Wert als negative Zahl
ws_filtered[f'E{text_start_row + 12}'].number_format = '#,##0.00 "€"'
ws_filtered[f'E{text_start_row + 12}'].font = Font(bold=True, color="000000")  # Fett formatieren
ws_filtered[f'E{text_start_row + 12}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Rahmen
ws_filtered[f'E{text_start_row + 12}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Hellblaue Füllung

# Spaltenbreite
ws_filtered.column_dimensions['A'].width = 12
ws_filtered.column_dimensions['B'].width = 35
ws_filtered.column_dimensions['C'].width = 12
ws_filtered.column_dimensions['D'].width = 12
ws_filtered.column_dimensions['E'].width = 12

# Speichern
wb_filtered.save(filtered_excel_path)

print(f"\n✓ Erfolgreich abgeschlossen!")
print(f"  {len(accounts)} Konten gefunden und exportiert")
print(f"  Excel-Datei (alle): {excel_path}")
print(f"  Excel-Datei (gefiltert): {filtered_excel_path} ({len(filtered_accounts)} Konten)")
